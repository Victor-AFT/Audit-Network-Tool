import telnetlib
from threading import Thread
import paramiko, os
from time import sleep
from paramiko import transport
from datetime import date
from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox, ttk
import ping3
from paramiko.ssh_exception import NoValidConnectionsError
import logging
import socket
import logging.handlers

import time,os
import datetime

current_date = datetime.date.today()

lista_usyscom=[]
lista_fortinet=[]
lista_ruggedcom=[]
lista_cisco=[]


transport.Transport._preferred_kex = (
    'ecdh-sha2-nistp256',
    'ecdh-sha2-nistp384',
    'ecdh-sha2-nistp521',
    'diffie-hellman-group-exchange-sha256',
    'diffie-hellman-group14-sha256',
    'diffie-hellman-group-exchange-sha1',
    'diffie-hellman-group14-sha1',
    'diffie-hellman-group1-sha1',
)



today = date.today()
fecha = today.strftime("%d_%m_%Y")
file_name_log="{0}.log".format(fecha)

def show_welcome_message():
    welcome_message = (
        "Bienvenido a la Aplicación de auditoria IP.\n"
        "#################################################.\n"
        "\n"
        "----------------------------------------------------------\n"
    )
    estado_label.config(text=welcome_message)

def schedule_check(t):
    root.after(1000, check_if_done, t)
def check_if_done(t):
    if not t.is_alive():
        a=0
    else:
        # Otherwise check again after one second.
        schedule_check(t)
def open_file():
    archivo_xls = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xls")])
    if archivo_xls and check_tyr.get()==0:
        mkdir(archivo_xls)
    if check_tyr.get()==1 and archivo_xls:
        check_session_tacacs_or_radius(archivo_xls)


def clear_messages():
    progress_bar["value"] = 0
    estado_label.config(text="")

def close_app():
    root.destroy()
def check_ping(host):
    response = ping3.ping(host)
    if response is not None:
        return True
    else:
        return False

def test_telnet(host,port):
    try:
        tn=telnetlib.Telnet(host,port,timeout=5)
        tn.close()
        return True
    except Exception as e:
        return False


def toggle_logging():

    if logging_var.get() == 1:
        logging.debug("Activado el modo log")
        logging.getLogger().disabled = False
        log_handler = logging.handlers.WatchedFileHandler(file_name_log)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        formatter.converter = time.gmtime  # if you want UTC time
        log_handler.setFormatter(formatter)
        logger = logging.getLogger()
        logger.addHandler(log_handler)
        logger.setLevel(logging.DEBUG)
        print("log activado")
    else:
        logging.getLogger().disabled = True
        logging.disable(logging.NOTSET)
        #c_handler = logging.handlers.WatchedFileHandler(file_name_log)
        #logging.getLogger().removeHandler(c_handler)
        #logger.removeHandler(c_handler)
        logging.debug("INFO","Se han desactivado los registros.")
        print("log desactivado")
def ConexionSSH_Ruggedcom(ip,user_ise,pssword_ise,demonicosubestacion):
    ruggedcom=[]
    ruggedcom.append('Ruggedcom')
    # print('Comprobando el ping con la direccion ip -> %s'%(ip))
    ruggedcom.append(ip)
    logging.debug('Comprobando conexión con Ruggedcom: %s', ip)
    if check_ping(ip):
        ruggedcom.append('up')
        logging.debug('Ping exitoso para Ruggedcom: %s', ip)
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            logging.debug('Intentando conexión SSH con Ruggedcom: %s', ip)
            'CONEXION SSH'
            client.connect(ip, 22, user_ise, pssword_ise, timeout=10, look_for_keys=False, allow_agent=False,
                           auth_timeout=15)
            shell = client.invoke_shell()
            sleep(1)
            shell.recv(1024)
            # Activates the menu
            shell.send(b'\n')  # ENTER
            sleep(1)
            shell.recv(1024)
            shell.send(b'\x13')  # CTRL+S ENTER SHELL
            sleep(1)
            shell.recv(1024)
            ################## SYSTEM ID #########################
            sleep(1)
            shell.send(bytes("sql select system Name from systemid \n", "utf-8"))
            sleep(2)
            config = str(shell.recv(10000).decode(encoding='utf-8')).split("\n")
            hostname_data = str(config).replace(" ", "").split(",")[3]
            hostname = hostname_data[3:-1]
            ruta = str(demonicosubestacion + "/" + str(hostname) + '_Ruggedcom.txt')
            # print(hostname)
            # print("Obteniendo informacion de %s - IP %s"%(hostname,ip))
            g = open(ruta, "w", encoding='latin-1')
            shell.send(b'\n')  # ENTER
            sleep(1)
            #################### VLAN PORT ###########################
            sleep(1)
            shell.send(bytes("sql select from vlanportcfg \n", "utf-8"))
            sleep(5)
            config = '\n'.join(shell.recv(102400).decode("utf-8").splitlines()[1:-3])
            # g = open(file_vlan, "w")
            g.write(str(config))
            # g.close()
            # print("IP: [%s] Obteniendo informacion de VLAN PORT"%(ip))
            shell.send(b'\n')  # ENTER
            sleep(1)
            #################### MAC ADDRESS ###########################
            shell.recv(1024)
            sleep(1)
            shell.send(bytes('sql select from macAddresses \n', "utf-8"))
            sleep(5)
            config = '\n'.join(shell.recv(102400).decode("utf-8").splitlines()[1:-3])
            # g = open(file_mac, "w")
            g.write(str(config))
            # g.close()
            # print("IP: [%s] Obteniendo informacion de MAC ADDRESS"%(ip))
            shell.send(b'\n')  # ENTER
            sleep(1)
            #################### ETH PORT CFG ###########################
            shell.recv(1024)
            sleep(1)
            shell.send(bytes('sql select from ethportcfg \n', "utf-8"))
            sleep(5)
            config = '\n'.join(shell.recv(102400).decode("utf-8").splitlines()[1:-3])
            # g = open(file_ethportcfg, "w")
            g.write(str(config))
            # g.close()
            # print("IP: [%s] Obteniendo informacion de ETHPORTCFG"%(ip))
            sleep(1)
            shell.send(b'\n')  # ENTER
            #################### ETH PORT STATUS ###########################
            sleep(1)
            shell.recv(1024)
            sleep(1)
            shell.send(bytes('sql select from ethportStatus \n', "utf-8"))
            sleep(5)
            config = '\n'.join(shell.recv(102400).decode("utf-8").splitlines()[1:-3])
            # print('ethportStatus')
            # print(config_ethportStatus)
            # g = open(file_ethportStatus, "w")
            g.write(str(config))
            # g.close()
            # print("IP: [%s] Obteniendo informacion de ETHPORTSTATUS"%(ip))
            sleep(1)
            ################## vlanStaticCfg ###############################
            sleep(1)
            shell.recv(1024)
            sleep(1)
            shell.send(bytes('sql select from vlanStaticCfg \n', "utf-8"))
            sleep(5)
            config = '\n'.join(shell.recv(102400).decode("utf-8").splitlines()[1:-3])
            # print('ethportStatus')
            # print(config_ethportStatus)
            g = open(file_vlanStaticCfg, "w")
            g.write(str(config))
            g.close()
            sleep(1)
            
            shell.send(b'\x18')
            client.close()
            g.close()

        except paramiko.AuthenticationException as auth_ex:
            client.close()
            ruggedcom.append('ssh_error')
            ruggedcom.append(auth_ex)
            logging.debug('Error de autenticación SSH para Ruggedcom %s: %s', ip, auth_ex)

        except paramiko.SSHException as ssh_ex:
            client.close()
            ruggedcom.append('ssh_error')
            ruggedcom.append(ssh_ex)
            logging.debug('Error SSH para Ruggedcom %s: %s', ip, ssh_ex)

        except NoValidConnectionsError as e:
            client.close()
            ruggedcom.append('ssh_error')
            ruggedcom.append(e)
            logging.debug('Error de conexión para Ruggedcom %s: %s', ip, e)
        except socket.error as se:
            if se.winerror==10060:
                client.close()
                ruggedcom.append('ssh_error')
                ruggedcom.append('Error durante el intento de conexión ya que el activo conectado no respondió 10060')
                logging.debug('Error de socket para Ruggedcom %s: %s', ip, se)


    else:
        ruggedcom.append('down')
        logging.debug('Ruggedcom %s está desconectado', ip)

    lista_ruggedcom.append(ruggedcom)
    return lista_ruggedcom


def check_SSH_Ruggedcom(ip,user_ise,pssword_ise):
    ruggedcom=[]
    ruggedcom.append('Ruggedcom')
    # print('Comprobando el ping con la direccion ip -> %s'%(ip))
    ruggedcom.append(ip)
    logging.debug('Comprobando conexión con Ruggedcom: %s', ip)
    if check_ping(ip):
        ruggedcom.append('up')
        logging.debug('Ping exitoso para Ruggedcom: %s', ip)
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            logging.debug('Intentando conexión SSH con Ruggedcom: %s', ip)
            'CONEXION SSH'
            client.connect(ip, 22, user_ise, pssword_ise, timeout=10, look_for_keys=False, allow_agent=False,
                           auth_timeout=15)
            ruggedcom.append('ssh_ok')
            client.close()

        except paramiko.AuthenticationException as auth_ex:
            client.close()
            ruggedcom.append('ssh_error')
            ruggedcom.append(auth_ex)
            logging.debug('Error de autenticación SSH para Ruggedcom %s: %s', ip, auth_ex)

        except paramiko.SSHException as ssh_ex:
            client.close()
            ruggedcom.append('ssh_error')
            ruggedcom.append(ssh_ex)
            logging.debug('Error SSH para Ruggedcom %s: %s', ip, ssh_ex)

        except NoValidConnectionsError as e:
            client.close()
            ruggedcom.append('ssh_error')
            ruggedcom.append(e)
            logging.debug('Error de conexión para Ruggedcom %s: %s', ip, e)
        except socket.error as se:
            if se.winerror==10060:
                client.close()
                ruggedcom.append('ssh_error')
                ruggedcom.append('Error durante el intento de conexión ya que el activo conectado no respondió 10060')
                logging.debug('Error de socket para Ruggedcom %s: %s', ip, se)

    else:
        ruggedcom.append('down')
        logging.debug('Ruggedcom %s está desconectado', ip)

    lista_ruggedcom.append(ruggedcom)
    return lista_ruggedcom

def ConexionSSH_Fortinet(ip,user_ise,pssword_ise,demonicosubestacion):

    fortinety=[]
    fortinety.append('Fortinet')
    fortinety.append(ip)
    logging.debug('Comprobando conexión con Fortinet: %s', str(ip))
    if check_ping(ip):
        fortinety.append('up')
        logging.debug('Ping exitoso para Fortinet: %s', str(ip))
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            logging.debug('Intentando conexión SSH con Fortinet: %s', str(ip))
            client.connect(hostname=ip, port=2322, username=user_ise, password=pssword_ise)
            connection = client.invoke_shell()
            ################### GET HOSTNAME #########################
            connection.send(b'')
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(1)
            config = str(connection.recv(10000).decode(encoding='latin-1').replace(" ", ""))
            nombre_subestacion = config.split("$")[0] + '_FORTINET.txt'
            ruta = str(demonicosubestacion + "/" + nombre_subestacion)
            g = open(ruta, "w", encoding='latin-1')
            # print("Obteniendo informacion de %s - IP %s" % (config.split("$")[0], ip))
            sleep(1)
            ################### GET SYS ARP #########################
            connection.send(b'get sys arp')
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(3)
            config = str(connection.recv(10000).decode(encoding='latin-1')).replace("  ", "").replace(
                "\n", "")
            # print(config)
            g.write(config)
            g.close()
            # print("IP: [%s] Obteniendo informacion GET SYS ARP" % (ip))
            connection.close()
            client.close()

        except paramiko.AuthenticationException as auth_ex:
            client.close()
            fortinety.append('ssh_error')
            fortinety.append(auth_ex)
            logging.debug('Error de autenticación SSH para Fortinet %s: %s', str(ip), str(auth_ex))

        except paramiko.SSHException as ssh_ex:
            client.close()
            fortinety.append('ssh_error')
            fortinety.append(ssh_ex)
            logging.debug('Error SSH para Fortinet %s: %s', str(ip), str(ssh_ex))
        except NoValidConnectionsError as e:
            client.close()
            fortinety.append('ssh_error')
            fortinety.append(e)
            logging.debug('Error de conexión para Fortinet %s: %s', str(ip), str(e))
        except socket.error as se:
            if se.winerror==10060:
                client.close()
                fortinety.append('ssh_error')
                fortinety.append('Error durante el intento de conexión ya que el activo conectado no respondió 10060')
                logging.debug('Error de socket para Fortinet %s: %s', str(ip), str(se))

    else:
        fortinety.append('down')
        logging.debug('Fortinet %s está desconectado', str(ip))

    lista_fortinet.append(fortinety)
    return lista_fortinet


def Check_SSH_Fortinet(ip,user_ise,pssword_ise):
    fortinety=[]
    fortinety.append('Fortinet')
    fortinety.append(ip)
    logging.debug('Comprobando conexión con Fortinet: %s', str(ip))
    if check_ping(ip):
        fortinety.append('up')
        logging.debug('Ping exitoso para Fortinet: %s', str(ip))
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            logging.debug('Intentando conexión SSH con Fortinet: %s', str(ip))
            client.connect(hostname=ip, port=2322, username=user_ise, password=pssword_ise)
            fortinety.append('ssh_ok')
            connection = client.invoke_shell()
            connection.close()
            client.close()

        except paramiko.AuthenticationException as auth_ex:
            client.close()
            fortinety.append('ssh_error')
            fortinety.append(auth_ex)
            logging.debug('Error de autenticación SSH para Fortinet %s: %s', str(ip), str(auth_ex))

        except paramiko.SSHException as ssh_ex:
            client.close()
            fortinety.append('ssh_error')
            fortinety.append(ssh_ex)
            logging.debug('Error SSH para Fortinet %s: %s', str(ip), str(ssh_ex))
        except NoValidConnectionsError as e:
            client.close()
            fortinety.append('ssh_error')
            fortinety.append(e)
            logging.debug('Error de conexión para Fortinet %s: %s', str(ip), str(e))
        except socket.error as se:
            if se.winerror==10060:
                client.close()
                fortinety.append('ssh_error')
                fortinety.append('Error durante el intento de conexión ya que el activo conectado no respondió 10060')
                logging.debug('Error de socket para Fortinet %s: %s', str(ip), str(se))

    else:
        fortinety.append('down')
        logging.debug('Fortinet %s está desconectado', str(ip))

    lista_fortinet.append(fortinety)
    return lista_fortinet


def ConexionSSH_usyscom(ip,user_ise,pssword_ise,demonicosubestacion):
    usyscom=[]
    usyscom.append('ZIV')
    usyscom.append(ip)
    logging.debug('Comprobando conexión con ZIV: %s', str(ip))
    if check_ping(ip):
        usyscom.append('up')
        logging.debug('Ping exitoso para ZIV: %s', str(ip))
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        try:
            logging.debug('Intentando conexión SSH con Ziv: %s', str(ip))
            client.connect(hostname=ip, port=22, username=user_ise, password=pssword_ise)
            connection = client.invoke_shell()
            ################### GET MAIN #########################
            connection.send(b'get main')
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(3)
            config = str(connection.recv(10000).decode(encoding='latin-1').splitlines()).replace("  ", '')
            hostname = ''
            if len(config.split(",")) == 27:
                hostname = config.split(",")[5].split('=')[1]
            sleep(5)
            nombre_subestacion = hostname.replace("'", "") + '_ZIV.txt'
            ruta = str(demonicosubestacion + "/" + nombre_subestacion)
            g = open(ruta, "w", encoding='latin-1')
            # print("Obteniendo informacion de %s - IP %s" % (hostname.replace("'", ""), ip))
            sleep(1)
            ################### GET PORT #########################
            connection.send(b'get port')
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(3)
            config = connection.recv(10000).decode(encoding='latin-1')
            g.write(config)
            # print("IP: [%s] Obteniendo informacion GET PORT" % (ip))
            sleep(5)
            ################### STATS PORT #########################
            connection.send(b'stats port')
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(5)
            config = connection.recv(10000).decode(encoding='latin-1')
            g.write(str(config))
            # print("IP: [%s] Obteniendo informacion STATS PORT" % (ip))
            sleep(10)
            ################### STATS MAC #########################
            connection.send(b'stats mac')
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(3)
            config = connection.recv(10000).decode(encoding='latin-1')
            g.write(str(config))
            # print("IP: [%s] Obteniendo informacion STATS MAC" % (ip))
            sleep(5)
            ################### GET VLAN #########################
            connection.send(b'get vlan')
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(1)
            connection.send(b'\n')  # ENTER
            sleep(3)
            config = connection.recv(10000).decode(encoding='latin-1')
            g.write(str(config))
            # print("IP: [%s] Obteniendo informacion GET VLAN" % (ip))
            sleep(1)
            connection.close()
            client.close()
            g.close()

        except paramiko.AuthenticationException as auth_ex:
            client.close()
            usyscom.append('ssh_error')
            usyscom.append(auth_ex)
            logging.debug('Error de autenticación SSH para Ziv %s: %s', str(ip), str(auth_ex))

        except paramiko.SSHException as ssh_ex:
            client.close()
            usyscom.append('ssh_error')
            usyscom.append(ssh_ex)
            logging.debug('Error SSH para Ziv %s: %s', str(ip), str(ssh_ex))
        except NoValidConnectionsError as e:
            client.close()
            usyscom.append('ssh_error')
            usyscom.append(e)
            logging.debug('Error de conexión para Ziv %s: %s', str(ip),str(e))

        except socket.error as se:
            if se.winerror==10060:
                client.close()
                usyscom.append('ssh_error')
                usyscom.append('Error durante el intento de conexión ya que el activo conectado no respondió 10060')
                logging.debug('Error de socket para Ziv %s: %s', str(ip), str(se))

    else:
        usyscom.append('down')
        logging.debug('Ziv %s está desconectado', str(ip))

    lista_usyscom.append(usyscom)
    return lista_usyscom

def Check_SSH_usyscom(ip,user_ise,pssword_ise,demonicosubestacion):
    usyscom=[]
    usyscom.append('ZIV')
    usyscom.append(ip)
    logging.debug('Comprobando conexión con ZIV: %s', str(ip))
    if check_ping(ip):
        usyscom.append('up')
        logging.debug('Ping exitoso para ZIV: %s', str(ip))
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        try:
            logging.debug('Intentando conexión SSH con Ziv: %s', str(ip))
            client.connect(hostname=ip, port=22, username=user_ise, password=pssword_ise)
            connection = client.invoke_shell()
            usyscom.append('ssh_ok')
            connection.close()
            client.close()

        except paramiko.AuthenticationException as auth_ex:
            client.close()
            usyscom.append('ssh_error')
            usyscom.append(auth_ex)
            logging.debug('Error de autenticación SSH para Ziv %s: %s', str(ip), str(auth_ex))

        except paramiko.SSHException as ssh_ex:
            client.close()
            usyscom.append('ssh_error')
            usyscom.append(ssh_ex)
            logging.debug('Error SSH para Ziv %s: %s', str(ip), str(ssh_ex))
        except NoValidConnectionsError as e:
            client.close()
            usyscom.append('ssh_error')
            usyscom.append(e)
            logging.debug('Error de conexión para Ziv %s: %s', str(ip),str(e))

        except socket.error as se:
            if se.winerror==10060:
                client.close()
                usyscom.append('ssh_error')
                usyscom.append('Error durante el intento de conexión ya que el activo conectado no respondió 10060')
                logging.debug('Error de socket para Ziv %s: %s', str(ip), str(se))

    else:
        usyscom.append('down')
        logging.debug('Ziv %s está desconectado', str(ip))

    lista_usyscom.append(usyscom)
    return lista_usyscom

def ConexionSSH_CISCO(ip,user_ise,pssword_ise,demonicosubestacion):
    cisco=[]
    cisco.append('CISCO')
    cisco.append(ip)
    logging.debug('Comprobando conexión con Cisco: %s', str(ip))

    if check_ping(ip):
        cisco.append('up')
        logging.debug('Ping exitoso para Cisco: %s', str(ip))
        true_telnet = 0
        if test_telnet(ip, 23):
            true_telnet = 1
            logging.debug('telnet activo')
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:

            # print("Proceso [%s] de [%s] " % (contador_cisco, len(list_ip)))
            logging.debug('Intentando conexión SSH con Ziv: %s', str(ip))
            client.connect(hostname=ip, port=22, username=user_ise, password=pssword_ise,timeout=10,auth_timeout=10)
            connection = client.invoke_shell()
            ################### HOSTNAME CISCO #########################
            connection.send(b"enable\n")
            sleep(1)
            connection.send(bytes(pssword_ise + '\n', 'utf-8'))  # ENTER
            sleep(1)
            config = connection.recv(10000).decode(encoding='latin-1')
            sleep(5)
            connection.send(b"sh run | include hostname\n")
            sleep(5)
            config = connection.recv(10000).decode(encoding='latin-1')
            hostname = config.split("\n")[1].split(" ")[1].replace("\r", "")
            # print("Obteniendo informacion de %s - IP %s" % (hostname, ip))
            ruta = str(demonicosubestacion + "/" + str(hostname) + '_CISCO.TXT')
            sleep(1)
            g = open(ruta, "w", encoding='latin-1')
            ################### GET PORT #########################
            sleep(1)
            connection.send(b"sh arp\n")
            sleep(1)
            connection.send(b'                 \n')
            sleep(5)
            g.write(connection.recv(10000).decode(encoding='latin-1'))
            sleep(1)
            ################### IP INTERFACE BRIEF #########################
            sleep(1)
            connection.send(b"sh ip interface brief\n")
            sleep(1)
            connection.send(b'  \n')
            sleep(5)
            g.write(connection.recv(10000).decode(encoding='latin-1'))
            sleep(1)
            ################### VLAN #########################
            sleep(1)
            connection.send(b"sh vlan\n")
            sleep(1)
            connection.send(b'\n')
            sleep(5)
            g.write(connection.recv(10000).decode(encoding='latin-1'))
            sleep(1)
            ################### INTERFACES DESCRIPTION#########################
            sleep(1)
            connection.send(b"sh interfaces description\n")
            sleep(1)
            connection.send(b'                       \n')
            sleep(5)
            g.write(connection.recv(10000).decode(encoding='latin-1'))
            sleep(5)
            ################### INTERFACES #########################
            sleep(1)
            connection.send(b"sh interfaces\n")
            sleep(1)
            connection.send(b'                       \n')
            sleep(5)
            g.write(connection.recv(10000).decode(encoding='latin-1'))
            sleep(1)
            connection.close()
            client.close()
            g.close()
        except paramiko.AuthenticationException as auth_ex:
            client.close()
            cisco.append('ssh_error')
            cisco.append(auth_ex)
            logging.error('Error de autenticación SSH para CISCO %s: %s', str(ip), str(auth_ex))

        except paramiko.SSHException as ssh_ex:
            client.close()
            cisco.append('ssh_error')
            cisco.append(ssh_ex)
            logging.error('Error SSH para CISCO %s: %s', str(ip), str(ssh_ex))

        except NoValidConnectionsError as e:
            client.close()
            cisco.append('ssh_error')
            cisco.append(e)
            logging.error('Error de conexión para CISCO %s: %s', str(ip), str(e))

        except socket.error as se:
            if se.winerror==10060:
                client.close()
                cisco.append('ssh_error')
                cisco.append('Error durante el intento de conexión ya que el activo conectado no respondió 10060')
                logging.error('Error de socket para CISCO %s: %s', str(ip), str(se))

        if true_telnet == 1:
            cisco.append("telnet_up")

    else:
        cisco.append("down")
        logging.debug('CISCO %s está desconectado', str(ip))
    lista_cisco.append(cisco)
    return lista_cisco


def Check_SSH_CISCO(ip,user_ise,pssword_ise):
    cisco=[]
    cisco.append('CISCO')
    cisco.append(ip)
    logging.debug('Comprobando conexión con Cisco: %s', str(ip))

    if check_ping(ip):
        cisco.append('up')
        logging.debug('Ping exitoso para Cisco: %s', str(ip))
        true_telnet = 0
        if test_telnet(ip, 23):
            true_telnet = 1
            logging.debug('telnet activo')
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:

            # print("Proceso [%s] de [%s] " % (contador_cisco, len(list_ip)))
            logging.debug('Intentando conexión SSH con CISCO: %s', str(ip))
            client.connect(hostname=ip, port=22, username=user_ise, password=pssword_ise,timeout=10,auth_timeout=10)
            cisco.append('ssh_ok')
            connection = client.invoke_shell()
            connection.close()
            client.close()

        except paramiko.AuthenticationException as auth_ex:
            client.close()
            cisco.append('ssh_error')
            cisco.append(auth_ex)
            logging.error('Error de autenticación SSH para Ziv %s: %s', str(ip), str(auth_ex))

        except paramiko.SSHException as ssh_ex:
            client.close()
            cisco.append('ssh_error')
            cisco.append(ssh_ex)
            logging.error('Error SSH para Ziv %s: %s', str(ip), str(ssh_ex))

        except NoValidConnectionsError as e:
            client.close()
            cisco.append('ssh_error')
            cisco.append(e)
            logging.error('Error de conexión para Ziv %s: %s', str(ip), str(e))

        except socket.error as se:
            if se.winerror==10060:
                client.close()
                cisco.append('ssh_error')
                cisco.append('Error durante el intento de conexión ya que el activo conectado no respondió 10060')
                logging.error('Error de socket para Ziv %s: %s', str(ip), str(se))

        if true_telnet == 1:
            cisco.append("telnet_up")

    else:
        cisco.append("down")
        logging.debug('Ziv %s está desconectado', str(ip))
    lista_cisco.append(cisco)
    return lista_cisco

def convert_xls_to_xlsx(archivo_xls):

    etiqueta["text"] = "Procesando informacion..."

    #boton_abrir['state']='disabled'

    archivo_xlsx = os.path.splitext(archivo_xls)[0] + ".xlsx"

    try:
        df = pd.read_excel(archivo_xls)
    except Exception as e:
        messagebox.showerror("Error", f"Error al leer el fabricante en el archivo xls: {e}")
        logging.debug("Error", f"Error al leer el fabricante en el archivo xls: {e}")
        return
    total_rows = df.shape[0]
    # Guardar en formato .xlsx
    try:
        with pd.ExcelWriter(archivo_xlsx, engine='openpyxl') as writer:
            for i in range(0, total_rows, 100):  # Simulando la conversión en partes
                df_part = df[i:i + 100]
                df_part.to_excel(writer, index=False)

    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar archivo xlsx: {e}")
        logging.debug("Error", f"Error al guardar archivo xlsx: {e}")
        return
    #print("excel xls convertido a xlsx",archivo_xlsx)
    logging.debug("Debug","excel: {0} convertido a -> {1} ".format(archivo_xlsx,archivo_xlsx))
    return archivo_xlsx

def contar_errores():
    errores = {
        'usyscom': {},
        'fortinet': {},
        'cisco': {},
        'ruggedcom': {}
    }

    for item in lista_usyscom:
        if 'ssh_error' in item:  # Verificar si el valor asociado con 'ssh_error' es verdadero
            error_type = item[-1]  # Suponiendo que el tipo de error está en el penúltimo elemento
            errores['usyscom'][error_type] = errores['usyscom'].get(error_type, 0) + 1

    for item in lista_fortinet:
        if 'ssh_error' in item:  # Verificar si el valor asociado con 'ssh_error' es verdadero
            error_type = item[-1]  # Suponiendo que el tipo de error está en el penúltimo elemento
            errores['fortinet'][error_type] = errores['fortinet'].get(error_type, 0) + 1

    for item in lista_cisco:
        if 'ssh_error' in item:  # Verificar si el valor asociado con 'ssh_error' es verdadero
            error_type = item[-1]  # Suponiendo que el tipo de error está en el penúltimo elemento
            errores['cisco'][error_type] = errores['cisco'].get(error_type, 0) + 1

    for item in lista_ruggedcom:
        if 'ssh_error' in item:  # Verificar si el valor asociado con 'ssh_error' es verdadero
            error_type = item[-1]  # Suponiendo que el tipo de error está en el penúltimo elemento
            errores['ruggedcom'][error_type] = errores['ruggedcom'].get(error_type, 0) + 1

    logging.debug("Errores encontrados en usyscom: %s", str(errores['usyscom']))
    logging.debug("Errores encontrados en fortinet: %s", str(errores['fortinet']))
    logging.debug("Errores encontrados en cisco: %s", str(errores['cisco']))
    logging.debug("Errores encontrados en ruggedcom: %s", str(errores['ruggedcom']))

    return errores


def contar_correctos():

    correcto = {
        'usyscom': {},
        'cisco': {},
        'ruggedcom': {}
    }

    for item in lista_usyscom:
        if 'ssh_ok' in item:  # Verificar si el valor asociado con 'ssh_error' es verdadero
            ok_type = item[-1]  # Suponiendo que el tipo de error está en el penúltimo elemento
            correcto['usyscom'][ok_type] = correcto['usyscom'].get(ok_type, 0) + 1

    for item in lista_cisco:
        if 'ssh_ok' in item:  # Verificar si el valor asociado con 'ssh_error' es verdadero
            ok_type = item[-1]  # Suponiendo que el tipo de error está en el penúltimo elemento
            correcto['cisco'][ok_type] = correcto['cisco'].get(ok_type, 0) + 1

    for item in lista_ruggedcom:
        if 'ssh_ok' in item:  # Verificar si el valor asociado con 'ssh_error' es verdadero
            ok_type = item[-1]  # Suponiendo que el tipo de error está en el penúltimo elemento
            correcto['ruggedcom'][ok_type] = correcto['ruggedcom'].get(ok_type, 0) + 1

    logging.debug("session correcta en usyscom: %s", str(correcto['usyscom']))
    #logging.debug("Errores encontrados en fortinet: %s", str(correcto['fortinet']))
    logging.debug("session correcta en cisco: %s", str(correcto['cisco']))
    logging.debug("session correcta en ruggedcom: %s", str(correcto['ruggedcom']))

    return correcto


def mostrar_errores():
    errores = contar_errores()
    logging.debug("Errores detectados durante la ejecución del programa: %s", str(errores))
    if errores:

        #logger.debug("Escribiendo lista de equipos con errores en el archivo 'equipos_con_errores.txt'")
        mensaje = "Se han encontrado los siguientes problemas:\n"
        """
        with open("equipos_con_errores.txt", "w") as f:
            f.write("Equipos con errores:\n")
            for lista, error_dict in errores.items():
                for error, _ in error_dict.items():
                    for item in globals()[f'lista_{lista}']:
                        f.write(f"{item[0]} - {item[1]} - {item[4]}\n")
        """
        for lista, error_dict in errores.items():
            for error, _ in error_dict.items():
                for item in globals()[f'lista_{lista}']:
                    mensaje += f"{item[0]} - {item[1]} - {item[4]}\n"
                # Puedes añadir más casos según los tipos de error que quieras detallar
        logging.debug("Errores detectados durante la ejecución del programa: %s", str(errores))
        return mensaje

def mostrar_correcto_errores():

    errores = contar_errores()
    correctos=contar_correctos()
    #logging.debug("Errores detectados durante la ejecución del programa: %s", str(errores))

    # logger.debug("Escribiendo lista de equipos con errores en el archivo 'equipos_con_errores.txt'")
    mensaje = "Informe de comprobacion:\n"
    if errores:
        for lista, error_dict in errores.items():
            for error, _ in error_dict.items():
                for item in globals()[f'lista_{lista}']:
                    mensaje += f"{item[0]} - {item[1]} - {item[4]}\n"
                # Puedes añadir más casos según los tipos de error que quieras detallar
    if correctos:
        for lista, correct_dict in correctos.items():
            for correct, _ in correct_dict.items():
                for item in globals()[f'lista_{lista}']:
                    mensaje += f"{item[0]} - {item[1]} - {item[4]}\n"
                # Puedes añadir más casos según los tipos de error que quieras detallar

    return mensaje


def mkdir(archivo_xls):

    file = simpledialog.askstring("Creacion de carpeta", "Introduce el nombre del directorio:",parent=root)

    #file='TEST'
    if os.path.exists(file):
        pass
        logging.debug("Existe el directorio ->",str(file))
    else:
        os.mkdir(file)
        logging.debug("Directorio creado: %s", str(file))
    if file:

        userr = ''
        pssword = ''
        user_ise = ''
        pssword_ise = ''

        try:
            wb = load_workbook(convert_xls_to_xlsx(archivo_xls))
            wb.active = 0
            ws = wb.active
            # print(wb.active)

            data_list = []
            'Pasamos la hoja al dictionario'
            headers = [cell.value for cell in ws[1]]
            # print(headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dit = {}
                for header, value in zip(headers, row):
                    row_dit[header] = value
                    data_list.append(row_dit)
            data_optz = []
            data_list_optz = []
            for s in data_list:
                data_optz = []
                for k, v in s.items():
                    if k == 'Caption':
                        data_optz.append(v)
                    if k == 'IP_Address':
                        data_optz.append(v)
                    if k == 'Vendor':
                        data_optz.append(v)
                data_list_optz.append(data_optz)

            data_limpio = []
            for i in data_list_optz:
                if i not in data_limpio:
                    data_limpio.append(i)
            'LISTA DE VENDEDORES'
            #print("LECTURA EXCEL XLSX: ",data_limpio)
            logging.debug("LECTURA EXCEL XLSX: {0} ".format(data_limpio))
            wb.close()
            ise = 0
            da = 0

            for x in data_limpio:
                if x[2] == 'RuggedCom Inc.' or x[2] == 'uSysCom':
                    ise = 1
                if x[2] == 'Fortinet, Inc.':
                    da = 1
            'CHECK SESION ISE'
            if ise == 1:
                user_ise = simpledialog.askstring("Autenticación ISE", "Introduce el usuario:", parent=root)
                logging.debug("Usuario ISE autenticado: %s", str(user_ise))


                if user_ise is None:
                    messagebox.showerror("Error Autenticación ISE", "Has cancelado la sesion del usuario ")
                    logging.debug("Error","Cancelada la autenticación de usuario ISE")
                    user_ise = simpledialog.askstring("Autenticación ISE", "Introduce el usuario:", parent=root)

                pssword_ise = simpledialog.askstring("Autenticación ISE", "Introduce la contraseña:", show='*',parent=root)


                if pssword_ise is  None:
                    messagebox.showerror("Error Autenticación ISE", "Has cancelado la introduccion de contraseña ")
                    pssword_ise = simpledialog.askstring("Autenticación ISE", "Introduce la contraseña:", show='*',parent=root)

            'CHECK SESION FORTINET'
            if da == 1:
                userr = simpledialog.askstring("Autenticación DA", "Introduce el usuario:",parent=root)
                logging.debug("Usuario DA autenticado: %s", str(userr))

                if userr is None:
                    messagebox.showerror("Error Autenticación DA", "Has cancelado la sesion del usuario ")
                    logging.debug("Error","Cancelada la autenticación de usuario DA")
                    userr = simpledialog.askstring("Autenticación DA", "Introduce el usuario:", parent=root)

                pssword = simpledialog.askstring("Autenticación DA", "Introduce la contraseña:", show='*',parent=root)

                if pssword is None:
                    messagebox.showerror("Error Autenticación DA", "Has cancelado la introduccion de contraseña")
                    pssword = simpledialog.askstring("Autenticación DA", "Introduce la contraseña:", parent=root)


            threads = []
            for x in data_limpio:
                if x[2] == 'RuggedCom Inc.':

                    tr = Thread(target=ConexionSSH_Ruggedcom, args=(x[1], user_ise, pssword_ise, file,))
                    tr.start()
                    schedule_check(tr)
                    threads.append(tr)

                if x[2] == 'Fortinet, Inc.':

                    tr = Thread(target=ConexionSSH_Fortinet, args=(x[1], userr, pssword, file,))
                    tr.start()
                    schedule_check(tr)
                    threads.append(tr)
                if x[2] == 'uSysCom':

                    tr = Thread(target=ConexionSSH_usyscom, args=(x[1], user_ise, pssword_ise, file,))
                    tr.start()
                    schedule_check(tr)
                    threads.append(tr)

                if x[2] == 'Cisco':

                    tr = Thread(target=ConexionSSH_CISCO, args=(x[1], user_ise, pssword_ise, file,))
                    tr.start()
                    schedule_check(tr)
                    threads.append(tr)

            for i in range(len(threads)):
                threads[i].join()
            progress_bar["value"] = len(threads)

            errores = mostrar_errores()
            if errores:

                progress_bar["value"] = 100
                etiqueta["text"] = "Proceso completado con errores"
                messagebox.showerror("Errores", errores)
                logging.debug("Errores mostrados a usuario: %s", str(errores))
            else:
                logging.debug("No hay errores detectados")
                messagebox.showinfo('Descarga Completada',
                                    'Se ha generado los siguientes ficheros: %s\n' % (os.listdir(file)), parent=root)
                etiqueta["text"] = "¡Archivo descargado correctamente!"
                progress_bar["value"] = 100

        except Exception as e:
            messagebox.showerror("Error","Error al leer excel "f"{e}")
            logging.debug("Error al leer el archivo Excel: %s", str(e))
    else:
        messagebox.showerror("Error","Directorio no creado..")
        logging.debug("Error","Directorio no creado..")


def check_session_tacacs_or_radius(archivo_xls):
    print("Se ha habilitado la comprobacion de sesión con tacacs y radius")
    logging.debug("Activado el modo comprobacion de usuario local o tacacas_y_radius")
    userr = ''
    pssword = ''
    user_ise = ''
    pssword_ise = ''

    try:
        wb = load_workbook(convert_xls_to_xlsx(archivo_xls))
        wb.active = 0
        ws = wb.active
        # print(wb.active)

        data_list = []
        'Pasamos la hoja al dictionario'
        headers = [cell.value for cell in ws[1]]
        # print(headers)
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dit = {}
            for header, value in zip(headers, row):
                row_dit[header] = value
                data_list.append(row_dit)
        data_optz = []
        data_list_optz = []
        for s in data_list:
            data_optz = []
            for k, v in s.items():
                if k == 'Caption':
                    data_optz.append(v)
                if k == 'IP_Address':
                    data_optz.append(v)
                if k == 'Vendor':
                    data_optz.append(v)
            data_list_optz.append(data_optz)

        data_limpio = []
        for i in data_list_optz:
            if i not in data_limpio:
                data_limpio.append(i)
        'LISTA DE VENDEDORES'
        # print("LECTURA EXCEL XLSX: ",data_limpio)
        logging.debug("LECTURA EXCEL XLSX: {0} ".format(data_limpio))
        wb.close()
        ise = 0
        da = 0

        for x in data_limpio:
            if x[2] == 'RuggedCom Inc.' or x[2] == 'uSysCom':
                ise = 1
            if x[2] == 'Fortinet, Inc.':
                da = 1
        'CHECK SESION ISE'
        if ise == 1:
            user_ise = simpledialog.askstring("Autenticación ISE", "Introduce el usuario:", parent=root)
            logging.debug("Usuario ISE autenticado: %s", str(user_ise))

            if user_ise is None:
                messagebox.showerror("Error Autenticación ISE", "Has cancelado la sesion del usuario ")
                logging.debug("Error", "Cancelada la autenticación de usuario ISE")
                user_ise = simpledialog.askstring("Autenticación ISE", "Introduce el usuario:", parent=root)

            pssword_ise = simpledialog.askstring("Autenticación ISE", "Introduce la contraseña:", show='*',
                                                 parent=root)

            if pssword_ise is None:
                messagebox.showerror("Error Autenticación ISE", "Has cancelado la introduccion de contraseña ")
                pssword_ise = simpledialog.askstring("Autenticación ISE", "Introduce la contraseña:", show='*',
                                                     parent=root)

        'CHECK SESION FORTINET'
        """
        if da == 1:
            userr = simpledialog.askstring("Autenticación DA", "Introduce el usuario:", parent=root)
            logging.debug("Usuario DA autenticado: %s", str(userr))

            if userr is None:
                messagebox.showerror("Error Autenticación DA", "Has cancelado la sesion del usuario ")
                logging.debug("Error", "Cancelada la autenticación de usuario DA")
                userr = simpledialog.askstring("Autenticación DA", "Introduce el usuario:", parent=root)

            pssword = simpledialog.askstring("Autenticación DA", "Introduce la contraseña:", show='*', parent=root)

            if pssword is None:
                messagebox.showerror("Error Autenticación DA", "Has cancelado la introduccion de contraseña")
                pssword = simpledialog.askstring("Autenticación DA", "Introduce la contraseña:", parent=root)
        """

        threads = []
        for x in data_limpio:
            if x[2] == 'RuggedCom Inc.':
                tr = Thread(target=check_SSH_Ruggedcom, args=(x[1], user_ise, pssword_ise,))
                tr.start()
                schedule_check(tr)
                threads.append(tr)
            """
            if x[2] == 'Fortinet, Inc.':
                tr = Thread(target=Check_SSH_Fortinet, args=(x[1], userr, pssword, ))
                tr.start()
                schedule_check(tr)
                threads.append(tr)
            """
            if x[2] == 'uSysCom':
                tr = Thread(target=Check_SSH_usyscom, args=(x[1], user_ise, pssword_ise, ))
                tr.start()
                schedule_check(tr)
                threads.append(tr)

            if x[2] == 'Cisco':
                tr = Thread(target=Check_SSH_CISCO, args=(x[1], user_ise, pssword_ise, ))
                tr.start()
                schedule_check(tr)
                threads.append(tr)

        for i in range(len(threads)):
            threads[i].join()
        progress_bar["value"] = len(threads)

        progress_bar["value"] = 100
        etiqueta["text"] = "Proceso completado"
        messagebox.showinfo("INFO", mostrar_correcto_errores())

    except Exception as e:
        messagebox.showerror("Error", "Error al leer excel "f"{e}")
        logging.debug("Error al leer el archivo Excel: %s", str(e))



# Configuración de la ventana principal
root = tk.Tk()
root.geometry('425x500+200+100')
root.title("Auditoria IP v6.1")


# Definición de estado_label antes de llamar a show_welcome_message
estado_label = tk.Label(text="")
estado_label.pack(pady=20)

show_welcome_message()

etiqueta = tk.Label(text="Por favor, selecciona un archivo .xls!")
etiqueta.pack(pady=10)
#etiqueta.place(x=200, y=200)
#etiqueta.config(width=200, height=200)

boton_abrir = tk.Button(text="Abrir", command=open_file)
boton_abrir.pack(pady=20)
boton_abrir.config(width=10, height=1)

style = ttk.Style(root)
style.theme_use('clam')
style.configure("blue.Horizontal.TProgressbar", background='blue', troughcolor='white')
progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate",style="blue.Horizontal.TProgressbar")
progress_bar.pack(pady=20)


boton_cerrar = tk.Button( text="Cerrar", command=close_app)
boton_cerrar.pack(pady=20)
boton_cerrar.config(width=10, height=1)


# Variable para almacenar el estado del cuadro de verificación
logging_var = tk.IntVar()
check_tyr=tk.IntVar()

# Crear el cuadro de verificación
logging_checkbox = ttk.Checkbutton(root, text="Activar log", variable=logging_var, command=toggle_logging)
logging_checkbox.pack(pady=10)

# Variable para almacenar el estado del cuadro de verificación
checktyr = tk.IntVar()
logging_checkbox = ttk.Checkbutton(root, text="Check tacacs+ y Radius", variable=check_tyr)
logging_checkbox.pack(pady=10)


root.mainloop()
